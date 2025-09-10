import asyncio
import asyncpg
import pandas as pd
import os
from datetime import date
from db import DB_URL
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side

COLUMNS = [
    "№ вылета",
    "Попадание в зону цели",
    "Попадание",
    "Объективный контроль",
    "Комментарий эксперта",
    "Вид цели",
    "Пилот",
    "В воздухе",
    "Время взлета",
    "Завершил полет",
    "Время завершения полета",
    "Уникальный пилот",
    "Кол-во взлетов у пилота",
    "Кол-во попаданий",
    "Кол-во промахов",
    "Кол-во незавершенных полетов"
]

GREEN = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
RED = PatternFill(start_color="FF7F7F", end_color="FF7F7F", fill_type="solid")
GREY = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

async def fetch_data():
    conn = await asyncpg.connect(DB_URL)
    rows = await conn.fetch("""
        SELECT t.id, u.squad, t.point, t.color, 
               t.true_point, t.true_color, 
               t.result, t.start_time, t.end_time, t.status
        FROM tasks t
        JOIN users u ON u.squad = t.squad
        WHERE t.status IN ('pending','accepted','finished')
          AND (t.created_at::date = CURRENT_DATE OR t.status IN ('pending','accepted'))
        ORDER BY t.id
    """)
    await conn.close()
    return rows

def transform(rows):
    data = []
    squad_stats = {}

    for r in rows:
        squad = r["squad"]
        if squad not in squad_stats:
            squad_stats[squad] = {"flights": 0, "hits": 0, "misses": 0, "unfinished": 0}

        if r["status"] in ("accepted", "finished"):
            squad_stats[squad]["flights"] += 1

        if r["status"] == "finished":
            if (r["true_point"] or (r["result"] and "Попадание" in r["result"])):
                squad_stats[squad]["hits"] += 1
            else:
                squad_stats[squad]["misses"] += 1
        elif r["status"] == "accepted":
            squad_stats[squad]["unfinished"] += 1

    squad_seen = set()
    for idx, r in enumerate(rows, start=1):
        hit_zone = 1 if r["true_point"] else 0
        hit = 1 if r["result"] and "Попадание" in r["result"] else 0
        target = f"{r['true_point']} {r['true_color']}" if r["true_point"] else f"{r['point']} {r['color']}"

        in_air, start, finished, end = "", "", "", ""
        if r["status"] == "accepted":
            in_air, start = "+", r["start_time"]
        elif r["status"] == "finished":
            in_air, start, finished, end = "x", r["start_time"], "+", r["end_time"]

        if r["squad"] not in squad_seen:
            uniq_pilot = r["squad"]
            stats = squad_stats[r["squad"]]
            flights, hits, misses, unfinished = stats["flights"], stats["hits"], stats["misses"], stats["unfinished"]
            squad_seen.add(r["squad"])
        else:
            uniq_pilot, flights, hits, misses, unfinished = "", "", "", "", ""

        data.append([
            idx,
            hit_zone,
            hit,
            0,
            "",
            target,
            r["squad"],
            in_air,
            start,
            finished,
            end,
            uniq_pilot,
            flights,
            hits,
            misses,
            unfinished
        ])
    return data

async def main():
    rows = await fetch_data()
    data = transform(rows)
    df = pd.DataFrame(data, columns=COLUMNS)

    # путь для сохранения
    os.makedirs("docx", exist_ok=True)
    today = date.today().strftime("%Y-%m-%d")
    filename = f"docx/daily_report_{today}.xlsx"

    df.to_excel(filename, index=False)

    wb = load_workbook(filename)
    ws = wb.active

    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # рамки и выравнивание
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    # перенос текста в заголовках
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # форматирование статусов
    for row in range(2, ws.max_row + 1):
        in_air, finished = ws[f"H{row}"].value, ws[f"J{row}"].value
        if in_air == "+": ws[f"H{row}"].fill = GREEN
        elif in_air == "x": ws[f"H{row}"].fill = RED
        if finished == "+": ws[f"J{row}"].fill = GREEN

    for row in range(2, ws.max_row + 1):
        if ws[f"N{row}"].value != "": ws[f"N{row}"].fill = GREEN
        if ws[f"O{row}"].value != "": ws[f"O{row}"].fill = RED
        if ws[f"P{row}"].value != "": ws[f"P{row}"].fill = GREY

    ws.auto_filter.ref = f"L1:P{ws.max_row}"

    # фиксированные ширины
    col_widths = {
        "A": 7.1, "B": 11.1, "C": 10.6, "D": 12.9,
        "E": 13.6, "F": 13.0, "G": 9.6, "H": 9.0,
        "I": 7.7, "J": 9.7, "K": 17.9, "L": 11.8,
        "M": 14.6, "N": 12.0, "O": 10.0, "P": 21.6,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # строка ИТОГО
    last_row = ws.max_row + 1
    ws[f"L{last_row}"] = "ИТОГО:"
    ws[f"M{last_row}"] = f"=SUM(M2:M{ws.max_row-1})"
    ws[f"N{last_row}"] = f"=SUM(N2:N{ws.max_row-1})"
    ws[f"O{last_row}"] = f"=SUM(O2:O{ws.max_row-1})"
    ws[f"P{last_row}"] = f"=SUM(P2:P{ws.max_row-1})"

    for col in ["L", "M", "N", "O", "P"]:
        cell = ws[f"{col}{last_row}"]
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        cell.fill = GREY

    wb.save(filename)
    print(f"✅ Report saved: {filename}")

# обработчик уведомлений
async def on_tasks_changed(conn, pid, channel, payload):
    print(f"🔔 Tasks table changed: {payload}")
    await main()

# слушатель PostgreSQL
async def listen_for_updates():
    conn = await asyncpg.connect(DB_URL)
    await conn.add_listener("tasks_changed", on_tasks_changed)
    print("👂 Listening for changes in tasks...")
    while True:
        await asyncio.sleep(3600)  # держим соединение открытым
